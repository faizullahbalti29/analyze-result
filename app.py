import json
import io
from collections import Counter
from statistics import mean
from flask import Flask, jsonify, request, send_file, render_template
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__, template_folder='templates', static_folder='static')

# ==========================================================
# LOAD DATA & PRE-PROCESS ON STARTUP
# ==========================================================
print("Loading students.json...")
try:
    with open("students.json", "r", encoding="utf-8") as f:
        students_data = json.load(f)
except FileNotFoundError:
    print("Error: students.json not found! Please run extract_students.py first.")
    students_data = []

# Group students by institution code
students_by_inst = {}
inst_names = {}
for s in students_data:
    code = s["institution_code"]
    name = s["institution"]
    if code not in students_by_inst:
        students_by_inst[code] = []
    students_by_inst[code].append(s)
    inst_names[code] = name

# Compile unique institutions list sorted by name
unique_institutions = sorted(
    [{"code": code, "name": name} for code, name in inst_names.items()],
    key=lambda x: x["name"]
)

print(f"Loaded {len(students_data)} students across {len(unique_institutions)} institutions.")

# ==========================================================
# HELPER STATS FUNCTIONS
# ==========================================================
def percentage(marks, total_marks=1100):
    return round((marks / total_marks) * 100, 2)

def calculate_summary(code, students):
    """Generate all statistics for one institution."""
    if not students:
        return None

    sorted_students = sorted(
        students,
        key=lambda x: x["marks"],
        reverse=True
    )
    
    marks = [s["marks"] for s in sorted_students]
    passed = sum(1 for s in sorted_students if s["status"] == "PASS")
    failed = len(sorted_students) - passed
    
    # Calculate grade counts
    grades = Counter(s["grade"] for s in sorted_students)
    
    summary = {
        "institution": inst_names.get(code, f"Institution {code}"),
        "code": code,
        "total_students": len(sorted_students),
        "passed": passed,
        "failed": failed,
        "pass_percentage": round((passed / len(sorted_students)) * 100, 2),
        "average_marks": round(mean(marks), 2) if marks else 0,
        "highest_marks": max(marks) if marks else 0,
        "lowest_marks": min(marks) if marks else 0,
        "topper": sorted_students[0] if sorted_students else None,
        "top10": sorted_students[:10],
        "grade_counts": {g: grades.get(g, 0) for g in ["A1", "A", "B", "C", "D", "E", "F"]}
    }
    return summary

# ==========================================================
# FLASK ROUTS / ENDPOINTS
# ==========================================================

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/institutions')
def get_institutions():
    """Return the list of all available institutions for autocomplete."""
    return jsonify(unique_institutions)

@app.route('/api/compare')
def compare():
    """Return metrics for the requested institution codes."""
    codes_str = request.args.get('codes', '')
    if not codes_str:
        return jsonify({"error": "No institution codes provided"}), 400
        
    codes = [c.strip() for c in codes_str.split(',') if c.strip()]
    
    results = {}
    for code in codes:
        inst_students = students_by_inst.get(code, [])
        if inst_students:
            summary = calculate_summary(code, inst_students)
            if summary:
                results[code] = summary
                
    if not results:
        return jsonify({"error": "None of the provided institution codes were found"}), 404
        
    # Sort leaderboard by topper marks
    leaderboard = sorted(
        results.values(),
        key=lambda x: x["topper"]["marks"] if x["topper"] else 0,
        reverse=True
    )
    
    # Overall merit list for compared institutions
    overall_merit = []
    for report in results.values():
        for student in report["top10"]:
            overall_merit.append({
                "institution": report["institution"],
                "institution_code": report["code"],
                **student
            })
    overall_merit.sort(key=lambda x: x["marks"], reverse=True)
    
    return jsonify({
        "reports": results,
        "leaderboard": leaderboard,
        "overall_merit": overall_merit
    })

@app.route('/api/export')
def export_excel():
    """Generate and export custom openpyxl Excel file for selected institutions."""
    codes_str = request.args.get('codes', '')
    total_marks = request.args.get('total_marks', default=1100, type=int)
    if not codes_str:
        return "No institution codes provided", 400
        
    codes = [c.strip() for c in codes_str.split(',') if c.strip()]
    
    reports = {}
    for code in codes:
        inst_students = students_by_inst.get(code, [])
        if inst_students:
            summary = calculate_summary(code, inst_students)
            if summary:
                reports[code] = summary
                
    if not reports:
        return "None of the provided institutions found", 404
        
    # Build Leaderboard
    leaderboard = sorted(
        reports.values(),
        key=lambda x: x["topper"]["marks"] if x["topper"] else 0,
        reverse=True
    )
    
    # Build Overall Merit
    overall_merit = []
    for report in reports.values():
        for student in report["top10"]:
            overall_merit.append({
                "institution": report["institution"],
                "institution_code": report["code"],
                **student
            })
    overall_merit.sort(key=lambda x: x["marks"], reverse=True)

    # Create Excel openpyxl Workbook
    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Styles
    HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    CENTER = Alignment(horizontal="center", vertical="center")
    
    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER

    def auto_width(ws):
        for column_cells in ws.columns:
            length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            ws.column_dimensions[column_cells[0].column_letter].width = max(length + 3, 10)

    def freeze_header(ws):
        ws.freeze_panes = "A2"

    # 1. DASHBOARD SHEET
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash.append(["FBISE GAZETTE ANALYSIS"])
    ws_dash.append([])
    ws_dash.append(["Exam", "SSC-II Annual Examination 2019"])
    ws_dash.append(["Compared Institutions", len(reports)])
    ws_dash.append(["Total Checked Students", sum(r["total_students"] for r in reports.values())])
    ws_dash.append([])
    ws_dash.append(["Institution", "Topper", "Marks", "Average", "Pass %", "Students"])
    style_header(ws_dash, 7)
    for report in leaderboard:
        ws_dash.append([
            report["institution"],
            report["topper"]["name"] if report["topper"] else "",
            report["topper"]["marks"] if report["topper"] else 0,
            report["average_marks"],
            report["pass_percentage"],
            report["total_students"]
        ])
    ws_dash["A1"].font = Font(size=18, bold=True)
    ws_dash.freeze_panes = "A8"
    auto_width(ws_dash)

    # 2. INSTITUTION SUMMARY SHEET
    ws_summary = wb.create_sheet("Institution Summary")
    ws_summary.append([
        "Institution", "Code", "Students", "Passed", "Failed", 
        "Pass %", "Average Marks", "Highest", "Lowest"
    ])
    style_header(ws_summary)
    for report in reports.values():
        ws_summary.append([
            report["institution"],
            report["code"],
            report["total_students"],
            report["passed"],
            report["failed"],
            report["pass_percentage"],
            report["average_marks"],
            report["highest_marks"],
            report["lowest_marks"]
        ])
    freeze_header(ws_summary)
    auto_width(ws_summary)

    # 3. INSTITUTION TOPPERS
    ws_toppers = wb.create_sheet("Institution Toppers")
    ws_toppers.append([
        "Rank", "Institution", "Code", "Roll No", 
        "Student Name", "Marks", "Percentage", "Grade", "Status"
    ])
    style_header(ws_toppers)
    for rank, report in enumerate(leaderboard, start=1):
        topper = report["topper"]
        if topper:
            ws_toppers.append([
                rank,
                report["institution"],
                report["code"],
                topper["roll_no"],
                topper["name"],
                topper["marks"],
                percentage(topper["marks"], total_marks),
                topper["grade"],
                topper["status"]
            ])
    freeze_header(ws_toppers)
    auto_width(ws_toppers)

    # 4. TOP 10 STUDENTS
    ws_top10 = wb.create_sheet("Top 10 Students")
    ws_top10.append([
        "Institution", "Institution Code", "Rank", "Roll No", 
        "Student Name", "Marks", "Percentage", "Grade", "Status"
    ])
    style_header(ws_top10)
    for report in reports.values():
        for rank, student in enumerate(report["top10"], start=1):
            ws_top10.append([
                report["institution"],
                report["code"],
                rank,
                student["roll_no"],
                student["name"],
                student["marks"],
                percentage(student["marks"], total_marks),
                student["grade"],
                student["status"]
            ])
        ws_top10.append([]) # Blank row separating institutions
    freeze_header(ws_top10)
    auto_width(ws_top10)

    # Save to memory stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Institution_Analysis_Selected.xlsx"
    )

if __name__ == '__main__':
    app.run(debug=True, port=5000)
