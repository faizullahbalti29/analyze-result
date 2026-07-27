import json
from collections import Counter
from statistics import mean
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ==========================================================
# CONFIGURATION
# ==========================================================

TOTAL_MARKS = 1100

TARGET_INSTITUTIONS = {
    "6676": "Public School and College, Sadpara Road, Skardu",
    "6705": "USWA Public School and College, Skardu (N.A), Gilgit-Baltistan",
    "6965": "Cadet College, Skardu",
}

OUTPUT_FILE = "Institution_Analysis_2019.xlsx"

# ==========================================================
# LOAD DATA
# ==========================================================

with open("target_institutions.json", "r", encoding="utf-8") as f:
    data = json.load(f)

# ==========================================================
# HELPER FUNCTIONS
# ==========================================================

def percentage(marks):
    """Return percentage from marks."""
    return round((marks / TOTAL_MARKS) * 100, 2)


def get_grade_counts(students):
    """Return grade-wise student counts."""
    grades = Counter()

    for student in students:
        grades[student["grade"]] += 1

    return grades


def calculate_summary(code, students):
    """Generate all statistics for one institution."""

    students = sorted(
        students,
        key=lambda x: x["marks"],
        reverse=True
    )

    marks = [s["marks"] for s in students]

    passed = sum(
        1
        for s in students
        if s["status"] == "PASS"
    )

    failed = len(students) - passed

    summary = {
        "institution": TARGET_INSTITUTIONS.get(code, code),
        "code": code,
        "students": students,

        "total_students": len(students),

        "passed": passed,
        "failed": failed,

        "pass_percentage": round(
            (passed / len(students)) * 100,
            2
        ),

        "average_marks": round(
            mean(marks),
            2
        ),

        "highest_marks": max(marks),

        "lowest_marks": min(marks),

        "topper": students[0],

        "top10": students[:10],

        "grade_counts": get_grade_counts(students)
    }

    return summary


# ==========================================================
# BUILD REPORT OBJECT
# ==========================================================

reports = {}

for code, students in data.items():

    reports[code] = calculate_summary(
        code,
        students
    )

print("=" * 70)
print("Institution Report Object Created")
print("=" * 70)

print(f"Institutions Loaded : {len(reports)}")

# ==========================================================
# OVERALL MERIT LIST
# ==========================================================

overall_merit = []

for report in reports.values():

    for student in report["students"]:

        overall_merit.append({

            "institution": report["institution"],

            "institution_code": report["code"],

            **student
        })

overall_merit.sort(
    key=lambda x: x["marks"],
    reverse=True
)

print(f"Overall Merit Students : {len(overall_merit)}")

# ==========================================================
# INSTITUTION LEADERBOARD
# ==========================================================

leaderboard = sorted(
    reports.values(),
    key=lambda x: x["topper"]["marks"],
    reverse=True
)

print("\nInstitution Leaderboard\n")

for rank, institution in enumerate(
    leaderboard,
    start=1
):

    topper = institution["topper"]

    print(
        f"{rank}. "
        f"{institution['institution']} | "
        f"{topper['name']} | "
        f"{topper['marks']}"
    )

# ==========================================================
# EXCEL WORKBOOK
# ==========================================================

wb = Workbook()

# Remove default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# ==========================================================
# COMMON EXCEL STYLES
# ==========================================================

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)

CENTER = Alignment(
    horizontal="center",
    vertical="center"
)

LEFT = Alignment(
    horizontal="left",
    vertical="center"
)


def style_header(ws, row=1):
    """
    Apply header formatting to any row.
    """

    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def auto_width(ws):
    """
    Automatically adjust column widths.
    """

    for column_cells in ws.columns:

        length = max(
            len(str(cell.value))
            if cell.value is not None
            else 0
            for cell in column_cells
        )

        ws.column_dimensions[
            column_cells[0].column_letter
        ].width = length + 3


def freeze_header(ws):
    """
    Freeze first row.
    """

    ws.freeze_panes = "A2"
# ==========================================================
# DASHBOARD SHEET
# ==========================================================

def create_dashboard():

    ws = wb.create_sheet("Dashboard")

    ws.append(["FBISE GAZETTE ANALYSIS"])
    ws.append([])
    ws.append(["Exam", "SSC-II Annual Examination 2019"])
    ws.append(["Compared Institutions", len(reports)])
    ws.append(["Total Students", len(overall_merit)])
    ws.append([])

    ws.append([
        "Institution",
        "Topper",
        "Marks",
        "Average",
        "Pass %",
        "Students"
    ])

    style_header(ws, 7)

    for report in leaderboard:

        ws.append([
            report["institution"],
            report["topper"]["name"],
            report["topper"]["marks"],
            report["average_marks"],
            report["pass_percentage"],
            report["total_students"]
        ])

    ws["A1"].font = Font(size=18, bold=True)

    freeze_header(ws)
    auto_width(ws)


# ==========================================================
# INSTITUTION SUMMARY SHEET
# ==========================================================

def create_summary_sheet():

    ws = wb.create_sheet("Institution Summary")

    headers = [
        "Institution",
        "Code",
        "Students",
        "Passed",
        "Failed",
        "Pass %",
        "Average Marks",
        "Highest",
        "Lowest"
    ]

    ws.append(headers)

    style_header(ws)

    for report in reports.values():

        ws.append([
            report["institution"],
            report["code"],
            report["total_students"],
            report["passed"],
            report["failed"],
            report["pass_percentage"],
            report["average_marks"],
            report["highest_marks"],
            report["lowest_marks"]
        ])

    freeze_header(ws)
    auto_width(ws)


# ==========================================================
# INSTITUTION TOPPERS
# ==========================================================

def create_toppers_sheet():

    ws = wb.create_sheet("Institution Toppers")

    headers = [
        "Rank",
        "Institution",
        "Code",
        "Roll No",
        "Student Name",
        "Marks",
        "Percentage",
        "Grade",
        "Status"
    ]

    ws.append(headers)
    style_header(ws)

    for rank, report in enumerate(leaderboard, start=1):

        topper = report["topper"]

        ws.append([
            rank,
            report["institution"],
            report["code"],
            topper["roll_no"],
            topper["name"],
            topper["marks"],
            percentage(topper["marks"]),
            topper["grade"],
            topper["status"]
        ])

    freeze_header(ws)
    ws.auto_filter.ref = ws.dimensions
    auto_width(ws)
    # ==========================================================
# TOP 10 STUDENTS
# ==========================================================

def create_top10_sheet():

    ws = wb.create_sheet("Top 10 Students")

    headers = [
        "Institution",
        "Institution Code",
        "Rank",
        "Roll No",
        "Student Name",
        "Marks",
        "Percentage",
        "Grade",
        "Status"
    ]

    ws.append(headers)
    style_header(ws)

    for report in reports.values():

        for rank, student in enumerate(report["top10"], start=1):

            ws.append([
                report["institution"],
                report["code"],
                rank,
                student["roll_no"],
                student["name"],
                student["marks"],
                percentage(student["marks"]),
                student["grade"],
                student["status"]
            ])

        # Blank row between institutions
        ws.append([])

    freeze_header(ws)
    ws.auto_filter.ref = ws.dimensions
    auto_width(ws)


# ==========================================================
# CREATE ALL SHEETS
# ==========================================================

create_dashboard()
create_summary_sheet()
create_toppers_sheet()
create_top10_sheet()

# ==========================================================
# SAVE WORKBOOK
# ==========================================================

wb.save(OUTPUT_FILE)
print(f"\nExcel file saved: {OUTPUT_FILE}")